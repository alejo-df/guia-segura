# users/views.py

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy
from django.contrib.auth.views import LoginView, PasswordResetView, PasswordChangeView
from django.contrib import messages
from django.contrib.messages.views import SuccessMessageMixin
from django.views import View
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout as django_logout
from django.db import models
from django.db.models import Q
import requests as req
import uuid
import os
import traceback
from datetime import datetime

# Modelos y Formularios
from django.views.generic import ListView
from django.contrib.auth.models import User
from django.contrib.auth.mixins import UserPassesTestMixin
from .forms import (
    FormularioRegistro,
    FormularioAcceso,
    FormularioActualizarUsuario,
    FormularioActualizarPerfil,
)
from .models import Perfil, HistorialGuia, ScrapingLog, HistorialNotificacion, IntentoLogin, AuditoriaUsuario

from django.contrib.auth.decorators import user_passes_test

#excel
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from django.core.mail import EmailMessage
from io import BytesIO
from django.conf import settings
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from .services import ejecutar_con_reintentos, enviar_estado_api_cliente


# ====================================================================
# 🔐 PERMISOS
# ====================================================================
def es_admin(user):
    return user.is_superuser or user.is_staff


def obtener_ip_cliente(request):
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


def registrar_auditoria(request=None, criterio="General", accion="", resultado="exitoso", detalle="", numero_guia="", usuario=None, username=""):
    """Registra eventos de usabilidad sin interrumpir el flujo si el log falla."""
    try:
        usuario_final = usuario
        username_final = username or ""
        ip = None
        metodo = ""
        ruta = ""

        if request is not None:
            ip = obtener_ip_cliente(request)
            metodo = request.method
            ruta = request.path[:255]
            if usuario_final is None and getattr(request, "user", None) and request.user.is_authenticated:
                usuario_final = request.user
            if not username_final:
                if usuario_final is not None:
                    username_final = usuario_final.username
                else:
                    username_final = request.POST.get("username", "") or request.GET.get("username", "") or "Anónimo"
        elif usuario_final is not None:
            username_final = usuario_final.username

        AuditoriaUsuario.objects.create(
            usuario=usuario_final if getattr(usuario_final, "is_authenticated", True) else None,
            username=username_final[:150],
            criterio=criterio[:100],
            accion=accion[:180],
            resultado=resultado,
            detalle=detalle,
            numero_guia=(numero_guia or "")[:200],
            ip=(ip or "")[:100],
            metodo=metodo[:10],
            ruta=ruta,
        )
    except Exception as e:
        print("Error registrando auditoría:", e)
        traceback.print_exc()


def nombre_rol(usuario):
    if usuario.is_superuser:
        return "Superusuario"
    if usuario.is_staff:
        return "Staff"
    return "Normal"


# ====================================================================
# 🔥 PANEL DE USUARIOS
# ====================================================================
@user_passes_test(es_admin)
def panel_usuarios(request):
    usuarios = User.objects.filter(is_active=True).order_by("id")
    return render(request, "users/panel_usuarios.html", {"usuarios": usuarios})


@user_passes_test(es_admin)
def usuarios_inactivos(request):
    usuarios = User.objects.filter(is_active=False).order_by("id")
    return render(request, "users/usuarios_inactivos.html", {"usuarios": usuarios})


@user_passes_test(es_admin)
def activar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    usuario.is_active = True
    usuario.save()
    registrar_auditoria(
        request,
        criterio="Usuarios",
        accion="Reactivó usuario",
        resultado="exitoso",
        detalle=f"El administrador reactivó la cuenta de {usuario.username}.",
        usuario=request.user,
    )
    messages.success(request, f"El usuario '{usuario.username}' fue reactivado.")
    return redirect("usuarios_inactivos")


@user_passes_test(es_admin)
def crear_usuario(request):
    if request.method == "POST":
        username = request.POST["username"]
        email = request.POST["email"]
        password = request.POST["password"]
        nuevo_usuario = User.objects.create_user(username=username, email=email, password=password)
        registrar_auditoria(
            request,
            criterio="Usuarios",
            accion="Creó usuario",
            resultado="exitoso",
            detalle=f"El administrador creó el usuario {nuevo_usuario.username} con correo {nuevo_usuario.email}.",
            usuario=request.user,
        )
        messages.success(request, "Usuario creado correctamente")
        return redirect("panel_usuarios")
    return render(request, "users/crear_usuario.html")


@user_passes_test(es_admin)
def editar_usuario(request, user_id):
    usuario = User.objects.get(id=user_id)
    if request.method == "POST":
        username_anterior = usuario.username
        email_anterior = usuario.email
        cambio_password = bool(request.POST["password"])
        usuario.username = request.POST["username"]
        usuario.email = request.POST["email"]
        if cambio_password:
            usuario.set_password(request.POST["password"])
        usuario.save()
        detalle = f"Usuario editado. Antes: {username_anterior} / {email_anterior}. Ahora: {usuario.username} / {usuario.email}."
        if cambio_password:
            detalle += " También se actualizó la contraseña."
        registrar_auditoria(
            request,
            criterio="Usuarios",
            accion="Editó usuario",
            resultado="exitoso",
            detalle=detalle,
            usuario=request.user,
        )
        messages.success(request, "Usuario actualizado")
        return redirect("panel_usuarios")
    return render(request, "users/editar_usuario.html", {"usuario": usuario})


@user_passes_test(es_admin)
def eliminar_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    if usuario == request.user:
        messages.error(request, "No puedes desactivar tu propia cuenta.")
        return redirect("panel_usuarios")
    usuario.is_active = False
    usuario.is_staff = False
    usuario.is_superuser = False
    usuario.save()
    registrar_auditoria(
        request,
        criterio="Usuarios",
        accion="Desactivó usuario",
        resultado="exitoso",
        detalle=f"El administrador desactivó la cuenta de {usuario.username}.",
        usuario=request.user,
    )
    messages.success(request, f"El usuario '{usuario.username}' fue desactivado correctamente.")
    return redirect("panel_usuarios")


@user_passes_test(es_admin)
def detalle_usuario(request, user_id):
    usuario = get_object_or_404(User, id=user_id)
    perfil = getattr(usuario, 'perfil', None)
    guias = (
        HistorialGuia.objects
        .filter(usuario=usuario, activo=True)
        .values("consulta_id", "numero_guia")
        .annotate(total_eventos=models.Count("id"), ultima=models.Max("fecha_consulta"))
        .order_by("-ultima")
    )
    return render(request, "users/detalle_usuario.html", {
        "usuario": usuario,
        "perfil": perfil,
        "guias": guias,
    })


# ====================================================================
# 🚀 CAMBIAR ROL DE USUARIO
# ====================================================================
@user_passes_test(es_admin)
def cambiar_rol_usuario(request, user_id, rol):
    usuario = get_object_or_404(User, id=user_id)
    if usuario == request.user:
        messages.error(request, "No puedes cambiar tu propio rol.")
        return redirect("panel_usuarios")
    rol_anterior = nombre_rol(usuario)
    if rol == "superuser":
        usuario.is_superuser = True
        usuario.is_staff = True
    elif rol == "staff":
        usuario.is_superuser = False
        usuario.is_staff = True
    elif rol == "normal":
        usuario.is_superuser = False
        usuario.is_staff = False
    else:
        registrar_auditoria(
            request,
            criterio="Roles",
            accion="Intentó cambiar rol",
            resultado="fallido",
            detalle=f"Rol inválido solicitado para {usuario.username}: {rol}.",
            usuario=request.user,
        )
        messages.error(request, "Rol inválido.")
        return redirect("panel_usuarios")
    usuario.save()
    rol_nuevo = nombre_rol(usuario)
    registrar_auditoria(
        request,
        criterio="Roles",
        accion="Cambió rol de usuario",
        resultado="exitoso",
        detalle=f"Usuario afectado: {usuario.username}. Rol anterior: {rol_anterior}. Rol nuevo: {rol_nuevo}.",
        usuario=request.user,
    )
    messages.success(request, f"El rol del usuario '{usuario.username}' fue actualizado a '{rol}'.")
    return redirect("panel_usuarios")


# ====================================================================
# HOME
# ====================================================================
def home(request):
    return render(request, 'users/home.html')


class VistaRegistro(View):
    form_class = FormularioRegistro
    initial = {'key': 'value'}
    template_name = 'users/register.html'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('/')
        return super().dispatch(request, *args, **kwargs)

    def get(self, request):
        form = self.form_class()
        return render(request, self.template_name, {'form': form})

    def post(self, request):
        form = self.form_class(request.POST)
        if form.is_valid():
            nuevo_usuario = form.save()
            registrar_auditoria(
                request,
                criterio="Registro",
                accion="Registro de usuario nuevo",
                resultado="exitoso",
                detalle=f"Se registró el usuario {nuevo_usuario.username} con correo {nuevo_usuario.email}.",
                usuario=nuevo_usuario,
                username=nuevo_usuario.username,
            )
            messages.success(request, "Cuenta creada")
            return redirect("login")
        return render(request, self.template_name, {'form': form})


@login_required
def profile(request):
    if request.method == 'POST':
        user_form = FormularioActualizarUsuario(request.POST, instance=request.user)
        profile_form = FormularioActualizarPerfil(request.POST, request.FILES, instance=request.user.perfil)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            registrar_auditoria(
                request,
                criterio="Perfil",
                accion="Actualizó perfil",
                resultado="exitoso",
                detalle="El usuario actualizó información de su perfil.",
            )
            messages.success(request, "Perfil actualizado")
            return redirect("users-profile")
    else:
        user_form = FormularioActualizarUsuario(instance=request.user)
        profile_form = FormularioActualizarPerfil(instance=request.user.perfil)

    contexto = {'user_form': user_form, 'profile_form': profile_form}

    if request.user.is_staff or request.user.is_superuser:
        return render(request, 'users/profile.html', contexto)
    else:
        return render(request, 'users/profile_usuario.html', contexto)


class VistaListaUsuarios(UserPassesTestMixin, ListView):
    model = User
    context_object_name = 'user_list'
    template_name = 'users/users_list.html'
    paginate_by = 15

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser


class VistaAccesoPersonalizada(LoginView):
    template_name = 'users/login.html'
    authentication_form = FormularioAcceso

    def form_valid(self, form):
        remember_me = form.cleaned_data.get('remember_me')
        if not remember_me:
            self.request.session.set_expiry(0)
        # Login exitoso: resetear intentos fallidos
        username = form.cleaned_data.get('username')
        usuario = form.get_user()
        try:
            IntentoLogin.objects.filter(username=username).update(
                intentos_fallidos=0, bloqueado_hasta=None
            )
        except Exception as e:
            print("Error actualizando intentos de login:", e)
            traceback.print_exc()

        registrar_auditoria(
            self.request,
            criterio="Autenticación",
            accion="Inicio de sesión",
            resultado="exitoso",
            detalle=f"El usuario {usuario.username} inició sesión correctamente.",
            usuario=usuario,
            username=usuario.username,
        )
        return super().form_valid(form)

    def form_invalid(self, form):
        # Login fallido: registrar intento
        username = self.request.POST.get('username', '')
        if username:
            bloqueado = False
            intentos_acumulados = "no disponible"
            try:
                from django.utils import timezone
                from datetime import timedelta
                intento, _ = IntentoLogin.objects.get_or_create(username=username)
                intento.intentos_fallidos += 1
                if intento.intentos_fallidos >= 5:
                    intento.bloqueado_hasta = timezone.now() + timedelta(minutes=15)
                    bloqueado = True
                intento.save()
                intentos_acumulados = intento.intentos_fallidos
            except Exception as e:
                print("Error registrando intento de login:", e)
                traceback.print_exc()

            registrar_auditoria(
                self.request,
                criterio="Autenticación",
                accion="Intento de inicio de sesión",
                resultado="bloqueado" if bloqueado else "fallido",
                detalle=f"Intento fallido para el usuario {username}. Intentos acumulados: {intentos_acumulados}.",
                username=username,
            )
        return super().form_invalid(form)


class VistaRestablecerContrasena(SuccessMessageMixin, PasswordResetView):
    template_name = 'users/password_reset.html'
    email_template_name = 'users/password_reset_email.html'
    subject_template_name = 'users/password_reset_subject.txt'
    success_message = "Se enviaron las instrucciones por correo"


class VistaCambiarContrasena(SuccessMessageMixin, PasswordChangeView):
    template_name = 'users/change_password.html'
    success_url = reverse_lazy('users-profile')
    success_message = "Contraseña cambiada"


@login_required
def cerrar_sesion(request):
    username = request.user.username
    registrar_auditoria(
        request,
        criterio="Autenticación",
        accion="Cierre de sesión",
        resultado="exitoso",
        detalle=f"El usuario {username} cerró sesión.",
    )
    django_logout(request)
    return redirect("login")


# ====================================================================
# 🔍 CONSULTAR GUÍA — via servidor local + ngrok
# ====================================================================
@login_required
def VistaConsultarGuia(request):
    resultados_consulta = None
    guia_consultada = None

    if request.method == 'POST':
        guia_consultada = request.POST.get('guia_a_consultar', '').strip()

        if guia_consultada:
            scraper_url = os.getenv('SCRAPER_URL', '').strip()
            scraper_secret = os.getenv('SCRAPER_SECRET', '').strip()

            if not scraper_url:
                registrar_auditoria(
                    request,
                    criterio="Consulta de guía",
                    accion="Consultó guía",
                    resultado="fallido",
                    detalle="El servidor de consultas no estaba activo o SCRAPER_URL no estaba configurado.",
                    numero_guia=guia_consultada,
                )
                messages.error(request, "El servidor de consultas no está activo. Asegúrate de tener ngrok corriendo.")
                return render(request, "users/consultar_guia.html", {
                    "resultados": None,
                    "guia_consultada": guia_consultada,
                })

            try:
                def consultar_scraper():
                    return req.get(
                        f"{scraper_url}/scrape",
                        params={"guia": guia_consultada},
                        headers={"X-API-Secret": scraper_secret},
                        timeout=60,
                    )

                response = ejecutar_con_reintentos(consultar_scraper, intentos=3, espera_inicial=2)

                if response.status_code == 200:
                    data = response.json()

                    if data.get("success") and data.get("eventos"):
                        eventos = data["eventos"]

                        ultimo = HistorialGuia.objects.aggregate(models.Max('consulta_id')).get("consulta_id__max")
                        nuevo_consulta_id = (ultimo or 0) + 1

                        for evento in eventos:
                            HistorialGuia.objects.create(
                                usuario=request.user,
                                consulta_id=nuevo_consulta_id,
                                numero_guia=guia_consultada,
                                fecha=evento["fecha"],
                                hora=evento["hora"],
                                estado=evento["estado"],
                                sucursal=evento["sucursal"],
                                fecha_consulta=datetime.now()
                            )

                        ultimo_evento = eventos[-1]
                        respuesta_cliente = enviar_estado_api_cliente({
                            "numero_guia": guia_consultada,
                            "consulta_id": nuevo_consulta_id,
                            "estado": ultimo_evento.get("estado"),
                            "fecha": ultimo_evento.get("fecha"),
                            "hora": ultimo_evento.get("hora"),
                            "sucursal": ultimo_evento.get("sucursal"),
                        })

                        HistorialNotificacion.objects.create(
                            numero_guia=guia_consultada,
                            canal="sistema",
                            destinatario=request.user.email or request.user.username,
                            mensaje=f"Consulta creada. Respuesta API cliente: {respuesta_cliente}",
                            enviado=bool(respuesta_cliente.get("success", True)),
                        )

                        registrar_auditoria(
                            request,
                            criterio="Consulta de guía",
                            accion="Consultó guía",
                            resultado="exitoso",
                            detalle=f"Consulta ID {nuevo_consulta_id}. Eventos obtenidos: {len(eventos)}. Último estado: {ultimo_evento.get('estado', '')}.",
                            numero_guia=guia_consultada,
                        )

                        resultados_consulta = eventos
                        messages.success(request, "Guía consultada correctamente")
                    else:
                        error_msg = data.get("error", "No se encontraron eventos para esta guía.")
                        ScrapingLog.objects.create(
                            numero_guia=guia_consultada,
                            tipo_error="sin_resultados",
                            mensaje=error_msg,
                        )
                        registrar_auditoria(
                            request,
                            criterio="Consulta de guía",
                            accion="Consultó guía",
                            resultado="fallido",
                            detalle=error_msg,
                            numero_guia=guia_consultada,
                        )
                        messages.warning(request, error_msg)
                else:
                    error_msg = f"Error al consultar la guía. Código: {response.status_code} - {response.text[:200]}"
                    ScrapingLog.objects.create(
                        numero_guia=guia_consultada,
                        tipo_error="http",
                        mensaje=error_msg,
                    )
                    registrar_auditoria(
                        request,
                        criterio="Consulta de guía",
                        accion="Consultó guía",
                        resultado="error",
                        detalle=error_msg,
                        numero_guia=guia_consultada,
                    )
                    messages.error(request, error_msg)

            except req.exceptions.Timeout:
                ScrapingLog.objects.create(
                    numero_guia=guia_consultada,
                    tipo_error="timeout",
                    mensaje="La consulta tardó demasiado (timeout 60s).",
                )
                registrar_auditoria(
                    request,
                    criterio="Consulta de guía",
                    accion="Consultó guía",
                    resultado="error",
                    detalle="La consulta tardó demasiado (timeout 60s).",
                    numero_guia=guia_consultada,
                )
                messages.error(request, "La consulta tardó demasiado. Intenta nuevamente.")
            except Exception as e:
                ScrapingLog.objects.create(
                    numero_guia=guia_consultada,
                    tipo_error="excepcion",
                    mensaje=str(e),
                )
                registrar_auditoria(
                    request,
                    criterio="Consulta de guía",
                    accion="Consultó guía",
                    resultado="error",
                    detalle=f"Error inesperado: {e}",
                    numero_guia=guia_consultada,
                )
                messages.error(request, f"Error inesperado: {e}")

    return render(request, "users/consultar_guia.html", {
        "resultados": resultados_consulta,
        "guia_consultada": guia_consultada,
    })


# ====================================================================
# 📦 PANEL DE GUÍAS
# ====================================================================
@user_passes_test(es_admin)
def panel_guias(request):
    registros = HistorialGuia.objects.filter(activo=True)

    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    estado = request.GET.get("estado")
    usuario = request.GET.get("usuario")

    if fecha_inicio:
        registros = registros.filter(fecha_consulta__date__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha_consulta__date__lte=fecha_fin)
    if estado:
        registros = registros.filter(estado__icontains=estado)
    if usuario:
        registros = registros.filter(usuario__username__icontains=usuario)

    consultas = (
        registros
        .values("consulta_id", "numero_guia", "usuario__username")
        .annotate(total_eventos=models.Count("id"))
        .order_by("-consulta_id")
    )

    return render(request, "users/guias_panel.html", {
        "consultas": consultas,
        "filtros": {
            "fecha_inicio": fecha_inicio or "",
            "fecha_fin": fecha_fin or "",
            "estado": estado or "",
            "usuario": usuario or "",
        }
    })


# ====================================================================
# 📄 DETALLE DE CONSULTA
# ====================================================================
@user_passes_test(es_admin)
def detalle_consulta(request, consulta_id):
    eventos = (
        HistorialGuia.objects
        .filter(consulta_id=consulta_id, activo=True)
        .order_by('id')
    )
    if not eventos.exists():
        messages.error(request, "No hay eventos activos para esta consulta.")
        return redirect("panel_guias")
    return render(request, "users/guias_detalle.html", {
        "consulta_id": consulta_id,
        "numero_guia": eventos.first().numero_guia,
        "eventos": eventos
    })


# ====================================================================
# ➕ CREAR EVENTO
# ====================================================================
@user_passes_test(es_admin)
def crear_evento(request, consulta_id):
    consulta = HistorialGuia.objects.filter(consulta_id=consulta_id).first()
    if not consulta:
        messages.error(request, "Consulta no encontrada.")
        return redirect("panel_guias")

    if request.method == "POST":
        estado = request.POST.get("estado")
        if estado == "otro":
            estado = request.POST.get("estado_otro")

        HistorialGuia.objects.create(
            usuario=request.user,
            consulta_id=consulta_id,
            numero_guia=consulta.numero_guia,
            fecha=request.POST.get("fecha"),
            hora=request.POST.get("hora"),
            estado=estado,
            sucursal=request.POST.get("sucursal"),
            fecha_consulta=datetime.now()
        )
        registrar_auditoria(
            request,
            criterio="Guías",
            accion="Creó evento de guía",
            resultado="exitoso",
            detalle=f"Consulta ID {consulta_id}. Estado registrado: {estado}.",
            numero_guia=consulta.numero_guia,
        )
        messages.success(request, "Evento creado correctamente.")
        return redirect("detalle_consulta", consulta_id=consulta_id)

    return render(request, "users/guias_crear_evento.html", {
        "consulta_id": consulta_id,
        "numero_guia": consulta.numero_guia,
    })


# ====================================================================
# ✏ EDITAR EVENTO
# ====================================================================
@user_passes_test(es_admin)
def editar_evento(request, consulta_id, evento_id):
    evento = get_object_or_404(HistorialGuia, id=evento_id)

    if request.method == "POST":
        estado = request.POST.get("estado")
        if estado == "otro":
            estado = request.POST.get("estado_otro")

        estado_anterior = evento.estado
        evento.fecha = request.POST.get("fecha")
        evento.hora = request.POST.get("hora")
        evento.estado = estado
        evento.sucursal = request.POST.get("sucursal")
        evento.save()

        registrar_auditoria(
            request,
            criterio="Guías",
            accion="Editó evento de guía",
            resultado="exitoso",
            detalle=f"Consulta ID {consulta_id}. Estado anterior: {estado_anterior}. Estado nuevo: {estado}.",
            numero_guia=evento.numero_guia,
        )
        messages.success(request, "Evento actualizado correctamente.")
        return redirect("detalle_consulta", consulta_id=consulta_id)

    return render(request, "users/guias_editar_evento.html", {
        "evento": evento,
        "consulta_id": consulta_id,
        "numero_guia": evento.numero_guia,
    })


# ====================================================================
# 🗑 ELIMINAR EVENTO (borrado lógico)
# ====================================================================
@user_passes_test(es_admin)
def eliminar_evento(request, consulta_id, evento_id):
    evento = get_object_or_404(HistorialGuia, id=evento_id)
    evento.activo = False
    evento.save()
    registrar_auditoria(
        request,
        criterio="Guías",
        accion="Eliminó evento de guía",
        resultado="exitoso",
        detalle=f"Consulta ID {consulta_id}. Evento ID {evento_id} marcado como inactivo.",
        numero_guia=evento.numero_guia,
    )
    messages.success(request, "Evento marcado como eliminado.")
    return redirect("detalle_consulta", consulta_id=consulta_id)


# ====================================================================
# 🔄 INACTIVOS Y RESTAURAR
# ====================================================================
@user_passes_test(es_admin)
def detalle_consulta_inactivos(request, consulta_id):
    eventos = HistorialGuia.objects.filter(consulta_id=consulta_id, activo=False).order_by("id")
    return render(request, "users/guias_detalle_inactivos.html", {
        "eventos": eventos,
        "consulta_id": consulta_id
    })


@user_passes_test(es_admin)
def restaurar_evento(request, consulta_id, evento_id):
    evento = get_object_or_404(HistorialGuia, id=evento_id)
    evento.activo = True
    evento.save()
    registrar_auditoria(
        request,
        criterio="Guías",
        accion="Restauró evento de guía",
        resultado="exitoso",
        detalle=f"Consulta ID {consulta_id}. Evento ID {evento_id} restaurado.",
        numero_guia=evento.numero_guia,
    )
    messages.success(request, "Evento restaurado correctamente.")
    return redirect("detalle_consulta_inactivos", consulta_id=consulta_id)


def obtener_registros_guias_filtrados(request):
    registros = HistorialGuia.objects.filter(activo=True)
    fecha_inicio = request.GET.get("fecha_inicio")
    fecha_fin = request.GET.get("fecha_fin")
    estado = request.GET.get("estado")
    usuario = request.GET.get("usuario")

    if fecha_inicio:
        registros = registros.filter(fecha_consulta__date__gte=fecha_inicio)
    if fecha_fin:
        registros = registros.filter(fecha_consulta__date__lte=fecha_fin)
    if estado:
        registros = registros.filter(estado__icontains=estado)
    if usuario:
        registros = registros.filter(usuario__username__icontains=usuario)

    return registros.order_by("numero_guia", "consulta_id", "fecha_consulta")


# ====================================================================
# 📊 EXPORTAR EXCEL
# ====================================================================
@user_passes_test(es_admin)
def exportar_guias_excel(request):
    registros = obtener_registros_guias_filtrados(request)

    if not registros.exists():
        registrar_auditoria(
            request,
            criterio="Reportes",
            accion="Generó reporte Excel",
            resultado="fallido",
            detalle="No había registros activos para exportar.",
        )
        messages.warning(request, "No hay registros activos para exportar.")
        return redirect("panel_guias")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Guías"

    columnas = [
        "ID Consulta", "Número de guía", "Usuario",
        "Fecha evento", "Hora", "Estado",
        "Sucursal / ciudad", "Fecha de consulta",
    ]

    for col_num, titulo in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=titulo)

    fila = 2
    for r in registros:
        ws.cell(row=fila, column=1, value=r.consulta_id)
        ws.cell(row=fila, column=2, value=r.numero_guia)
        ws.cell(row=fila, column=3, value=r.usuario.username if r.usuario else "")
        ws.cell(row=fila, column=4, value=r.fecha)
        ws.cell(row=fila, column=5, value=r.hora)
        ws.cell(row=fila, column=6, value=r.estado)
        ws.cell(row=fila, column=7, value=r.sucursal)
        ws.cell(row=fila, column=8, value=r.fecha_consulta.strftime("%Y-%m-%d %H:%M:%S") if r.fecha_consulta else "")
        fila += 1

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.getvalue()

    if request.user.email:
        try:
            mensaje = EmailMessage(
                subject="Reporte de guías generado",
                body=f"Hola {request.user.username},\n\nAdjunto encontrarás el reporte de guías.",
                from_email=settings.EMAIL_HOST_USER,
                to=[request.user.email],
            )
            mensaje.attach(
                "reporte_guias.xlsx",
                excel_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            mensaje.send()
            HistorialNotificacion.objects.create(
                numero_guia="REPORTE",
                canal="email",
                destinatario=request.user.email,
                mensaje="Reporte de guías Excel enviado al correo del usuario.",
                enviado=True,
            )
            messages.success(request, f"Reporte enviado a {request.user.email}")
        except Exception as e:
            messages.error(request, f"Error enviando correo: {e}")
    else:
        messages.warning(request, "El usuario no tiene correo configurado.")

    response = HttpResponse(
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="reporte_guias.xlsx"'
    registrar_auditoria(
        request,
        criterio="Reportes",
        accion="Generó reporte Excel",
        resultado="exitoso",
        detalle=f"Reporte Excel administrativo generado con {registros.count()} registros.",
    )
    return response


# ====================================================================
# 📄 EXPORTAR PDF
# ====================================================================
@user_passes_test(es_admin)
def exportar_guias_pdf(request):
    registros = obtener_registros_guias_filtrados(request)

    if not registros.exists():
        registrar_auditoria(
            request,
            criterio="Reportes",
            accion="Generó reporte PDF",
            resultado="fallido",
            detalle="No había registros activos para exportar.",
        )
        messages.warning(request, "No hay registros activos para exportar.")
        return redirect("panel_guias")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="reporte_guias.pdf"'

    pdf = canvas.Canvas(response, pagesize=letter)
    ancho, alto = letter
    y = alto - 50

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, y, "Reporte consolidado de guías")
    y -= 30
    pdf.setFont("Helvetica", 8)

    encabezado = "ID | Guía | Usuario | Fecha evento | Hora | Estado | Sucursal | Fecha consulta"
    pdf.drawString(40, y, encabezado)
    y -= 15

    for r in registros:
        linea = f"{r.consulta_id} | {r.numero_guia} | {r.usuario.username if r.usuario else ''} | {r.fecha or ''} | {r.hora or ''} | {r.estado or ''} | {r.sucursal or ''} | {r.fecha_consulta.strftime('%Y-%m-%d %H:%M') if r.fecha_consulta else ''}"
        pdf.drawString(40, y, linea[:150])
        y -= 14
        if y < 50:
            pdf.showPage()
            pdf.setFont("Helvetica", 8)
            y = alto - 50

    pdf.save()
    registrar_auditoria(
        request,
        criterio="Reportes",
        accion="Generó reporte PDF",
        resultado="exitoso",
        detalle=f"Reporte PDF administrativo generado con {registros.count()} registros.",
    )
    return response


# ====================================================================
# 🧾 HISTORIALES ADMINISTRATIVOS - SPRINT 4
# ====================================================================


@user_passes_test(es_admin)
def panel_auditoria_usuarios(request):
    logs = AuditoriaUsuario.objects.select_related("usuario").all()

    q = request.GET.get("q", "").strip()
    usuario = request.GET.get("usuario", "").strip()
    criterio = request.GET.get("criterio", "").strip()
    accion = request.GET.get("accion", "").strip()
    resultado = request.GET.get("resultado", "").strip()
    fecha_inicio = request.GET.get("fecha_inicio", "").strip()
    fecha_fin = request.GET.get("fecha_fin", "").strip()

    if q:
        logs = logs.filter(
            Q(username__icontains=q) |
            Q(criterio__icontains=q) |
            Q(accion__icontains=q) |
            Q(detalle__icontains=q) |
            Q(numero_guia__icontains=q) |
            Q(ruta__icontains=q)
        )
    if usuario:
        logs = logs.filter(username__icontains=usuario)
    if criterio:
        logs = logs.filter(criterio__icontains=criterio)
    if accion:
        logs = logs.filter(accion__icontains=accion)
    if resultado:
        logs = logs.filter(resultado=resultado)
    if fecha_inicio:
        logs = logs.filter(fecha__date__gte=fecha_inicio)
    if fecha_fin:
        logs = logs.filter(fecha__date__lte=fecha_fin)

    criterios = AuditoriaUsuario.objects.values_list("criterio", flat=True).distinct().order_by("criterio")
    acciones = AuditoriaUsuario.objects.values_list("accion", flat=True).distinct().order_by("accion")

    return render(request, "users/panel_auditoria_usuarios.html", {
        "logs": logs[:300],
        "criterios": criterios,
        "acciones": acciones,
        "filtros": {
            "q": q,
            "usuario": usuario,
            "criterio": criterio,
            "accion": accion,
            "resultado": resultado,
            "fecha_inicio": fecha_inicio,
            "fecha_fin": fecha_fin,
        }
    })


@user_passes_test(es_admin)
def panel_logs_scraping(request):
    logs = ScrapingLog.objects.order_by("-fecha")[:100]
    return render(request, "users/panel_logs_scraping.html", {"logs": logs})


@user_passes_test(es_admin)
def panel_notificaciones(request):
    notificaciones = HistorialNotificacion.objects.order_by("-fecha")[:100]
    return render(request, "users/panel_notificaciones.html", {"notificaciones": notificaciones})


@user_passes_test(es_admin)
def panel_intentos_login(request):
    intentos = IntentoLogin.objects.order_by("-ultimo_intento")[:100]
    return render(request, "users/panel_intentos_login.html", {"intentos": intentos})


@user_passes_test(es_admin)
def desbloquear_cuenta(request, intento_id):
    intento = get_object_or_404(IntentoLogin, id=intento_id)
    intento.intentos_fallidos = 0
    intento.bloqueado_hasta = None
    intento.save()
    messages.success(request, f"Cuenta '{intento.username}' desbloqueada.")
    return redirect("panel_intentos_login")


# ====================================================================
# 📊 EXPORTAR GUÍAS DEL USUARIO (Excel y PDF)
# ====================================================================
@login_required
def mis_guias_excel(request):
    registros = HistorialGuia.objects.filter(
        usuario=request.user, activo=True
    ).order_by("consulta_id", "fecha_consulta")

    if not registros.exists():
        registrar_auditoria(
            request,
            criterio="Reportes",
            accion="Descargó reporte Excel personal",
            resultado="fallido",
            detalle="El usuario no tenía guías guardadas para exportar.",
        )
        messages.warning(request, "No tienes guías guardadas para exportar.")
        return redirect("users-profile")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mis Guías"

    columnas = ["ID Consulta", "Número de Guía", "Fecha evento", "Hora", "Estado", "Sucursal", "Fecha consulta"]
    for col_num, titulo in enumerate(columnas, 1):
        ws.cell(row=1, column=col_num, value=titulo)

    for fila, r in enumerate(registros, 2):
        ws.cell(row=fila, column=1, value=r.consulta_id)
        ws.cell(row=fila, column=2, value=r.numero_guia)
        ws.cell(row=fila, column=3, value=r.fecha)
        ws.cell(row=fila, column=4, value=r.hora)
        ws.cell(row=fila, column=5, value=r.estado)
        ws.cell(row=fila, column=6, value=r.sucursal)
        ws.cell(row=fila, column=7, value=r.fecha_consulta.strftime("%Y-%m-%d %H:%M:%S") if r.fecha_consulta else "")

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    excel_bytes = buffer.getvalue()

    # Registrar notificación
    HistorialNotificacion.objects.create(
        numero_guia="REPORTE_PROPIO",
        canal="sistema",
        destinatario=request.user.username,
        mensaje=f"El usuario descargó su reporte Excel personal.",
        enviado=True,
    )

    response = HttpResponse(
        excel_bytes,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="mis_guias_{request.user.username}.xlsx"'
    registrar_auditoria(
        request,
        criterio="Reportes",
        accion="Descargó reporte Excel personal",
        resultado="exitoso",
        detalle=f"El usuario descargó su reporte Excel personal con {registros.count()} registros.",
    )
    return response


@login_required
def mis_guias_pdf(request):
    registros = HistorialGuia.objects.filter(
        usuario=request.user, activo=True
    ).order_by("consulta_id", "fecha_consulta")

    if not registros.exists():
        registrar_auditoria(
            request,
            criterio="Reportes",
            accion="Descargó reporte PDF personal",
            resultado="fallido",
            detalle="El usuario no tenía guías guardadas para exportar.",
        )
        messages.warning(request, "No tienes guías guardadas para exportar.")
        return redirect("users-profile")

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="mis_guias_{request.user.username}.pdf"'

    pdf = canvas.Canvas(response, pagesize=letter)
    ancho, alto = letter
    y = alto - 50

    pdf.setFont("Helvetica-Bold", 14)
    pdf.drawString(40, y, f"Mis guías — {request.user.username}")
    y -= 30
    pdf.setFont("Helvetica", 8)
    pdf.drawString(40, y, "ID | Guía | Fecha | Hora | Estado | Sucursal")
    y -= 15

    for r in registros:
        linea = f"{r.consulta_id} | {r.numero_guia} | {r.fecha or ''} | {r.hora or ''} | {r.estado or ''} | {r.sucursal or ''}"
        pdf.drawString(40, y, linea[:140])
        y -= 14
        if y < 50:
            pdf.showPage()
            pdf.setFont("Helvetica", 8)
            y = alto - 50

    pdf.save()

    HistorialNotificacion.objects.create(
        numero_guia="REPORTE_PROPIO",
        canal="sistema",
        destinatario=request.user.username,
        mensaje=f"El usuario descargó su reporte PDF personal.",
        enviado=True,
    )

    registrar_auditoria(
        request,
        criterio="Reportes",
        accion="Descargó reporte PDF personal",
        resultado="exitoso",
        detalle=f"El usuario descargó su reporte PDF personal con {registros.count()} registros.",
    )

    return response


@login_required
def mis_notificaciones(request):
    notificaciones = HistorialNotificacion.objects.filter(
        destinatario__in=[request.user.email, request.user.username]
    ).order_by("-fecha")[:50]
    return render(request, "users/mis_notificaciones.html", {
        "notificaciones": notificaciones,
    })